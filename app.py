import streamlit as st

st.set_page_config(
    page_title="Calculadora de Frete",
    page_icon="🚛",
    layout="centered",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

.stApp {
    background-color: #0f1117;
    color: #e8e6e0;
}

h1, h2, h3 {
    font-family: 'IBM Plex Mono', monospace !important;
    letter-spacing: -0.02em;
}

.block-container {
    padding-top: 2rem;
    max-width: 720px;
}

.result-card {
    background: #1a1d27;
    border: 1px solid #2e3347;
    border-radius: 10px;
    padding: 20px 24px;
    margin-bottom: 16px;
}

.result-label {
    font-size: 12px;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 4px;
    font-family: 'IBM Plex Mono', monospace;
}

.result-value {
    font-size: 28px;
    font-weight: 600;
    font-family: 'IBM Plex Mono', monospace;
    letter-spacing: -0.02em;
}

.val-green { color: #4ade80; }
.val-red   { color: #f87171; }
.val-blue  { color: #60a5fa; }
.val-amber { color: #fbbf24; }

.formula-box {
    background: #12141d;
    border-left: 3px solid #3b82f6;
    border-radius: 0 6px 6px 0;
    padding: 12px 16px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 13px;
    color: #9ca3af;
    margin-top: 12px;
    line-height: 1.8;
}

.rota-badge {
    display: inline-block;
    background: #1e2433;
    border: 1px solid #2e3347;
    border-radius: 6px;
    padding: 4px 10px;
    font-size: 12px;
    font-family: 'IBM Plex Mono', monospace;
    color: #9ca3af;
    margin: 2px;
}

.section-header {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #4b5563;
    margin-bottom: 12px;
    margin-top: 8px;
    border-bottom: 1px solid #1e2433;
    padding-bottom: 6px;
}

div[data-testid="stNumberInput"] input,
div[data-testid="stTextInput"] input,
div[data-testid="stSelectbox"] select {
    background: #1a1d27 !important;
    border: 1px solid #2e3347 !important;
    color: #e8e6e0 !important;
    border-radius: 6px !important;
    font-family: 'IBM Plex Mono', monospace !important;
}

div[data-testid="stCheckbox"] label {
    font-size: 14px;
    color: #c9c7bf;
}

.stSelectbox label, .stNumberInput label, .stTextInput label {
    color: #6b7280 !important;
    font-size: 12px !important;
    font-family: 'IBM Plex Mono', monospace !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
}

div[data-testid="stButton"] button {
    background: #1e2433;
    border: 1px solid #2e3347;
    color: #9ca3af;
    border-radius: 6px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 12px;
    width: 100%;
}

div[data-testid="stButton"] button:hover {
    background: #252a3d;
    border-color: #3b82f6;
    color: #60a5fa;
}

.stDownloadButton button {
    background: #1a2e1a !important;
    border: 1px solid #166534 !important;
    color: #4ade80 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 13px !important;
    width: 100% !important;
    border-radius: 6px !important;
}
</style>
""", unsafe_allow_html=True)

ROTAS = {
    "JP / Fortaleza":        {"frete": 4500, "comissao": 600},
    "Fortaleza / Recife":    {"frete": 4500, "comissao": 600},
    "Recife / Natal":        {"frete": 3500, "comissao": 400},
    "Rota longa (R$5.700)":  {"frete": 5700, "comissao": 600},
    "Fort / Rec (R$4.860)":  {"frete": 4860, "comissao": 600},
    "JP / Teresina":         {"frete": 11000, "comissao": 1200},
    "Curta (R$2.500)":       {"frete": 2500, "comissao": 200},
    "Personalizada":         {"frete": None, "comissao": None},
}

st.markdown("# 🚛 Calculadora de Frete")
st.markdown("<div style='color:#4b5563;font-family:IBM Plex Mono,monospace;font-size:13px;margin-bottom:24px'>Cálculo de pagamento por viagem</div>", unsafe_allow_html=True)

col_mot, col_data = st.columns(2)
with col_mot:
    motorista = st.text_input("Motorista", placeholder="ex: Edivaldo")
with col_data:
    import datetime
    data_viagem = st.date_input("Data da viagem", value=datetime.date.today())

st.markdown("<div class='section-header'>Rota</div>", unsafe_allow_html=True)
rota_selecionada = st.selectbox("Selecione a rota", list(ROTAS.keys()), label_visibility="collapsed")

defaults = ROTAS[rota_selecionada]

st.markdown("<div class='section-header'>Valores</div>", unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    frete = st.number_input("Frete (R$)", min_value=0, step=100,
                            value=defaults["frete"] if defaults["frete"] else 4500)
with col2:
    pct_adi = st.number_input("Adiantamento (%)", min_value=0, max_value=100,
                               step=1, value=70)
with col3:
    abastecimento = st.number_input("Abastecimento (R$)", min_value=0, step=50, value=0)

st.markdown("<div class='section-header'>Deduções do saldo</div>", unsafe_allow_html=True)

col_a, col_b = st.columns(2)
with col_a:
    usar_comissao = st.checkbox("Comissão", value=True)
    comissao = st.number_input("Valor comissão (R$)", min_value=0, step=50,
                                value=defaults["comissao"] if defaults["comissao"] else 600,
                                disabled=not usar_comissao)

    usar_gerente = st.checkbox("Gerente", value=False)
    gerente = st.number_input("Valor gerente (R$)", min_value=0, step=50,
                               value=200, disabled=not usar_gerente)

with col_b:
    usar_guarda = st.checkbox("Guarda", value=True)
    guarda = st.number_input("Valor guarda (R$)", min_value=0, step=5,
                              value=20, disabled=not usar_guarda)

    usar_pneu = st.checkbox("Pneu / borracharia", value=False)
    pneu = st.number_input("Valor pneu/borr. (R$)", min_value=0, step=50,
                            value=200, disabled=not usar_pneu)

usar_outros = st.checkbox("Outros", value=False)
outros_desc = st.text_input("Descrição outros", value="", disabled=not usar_outros, placeholder="ex: serv. diversos")
outros = st.number_input("Valor outros (R$)", min_value=0, step=10, value=0, disabled=not usar_outros)

adiantamento = frete * pct_adi / 100
saldo_bruto  = frete - adiantamento

total_deducoes = (
    (comissao if usar_comissao else 0) +
    (guarda   if usar_guarda   else 0) +
    (gerente  if usar_gerente  else 0) +
    (pneu     if usar_pneu     else 0) +
    (outros   if usar_outros   else 0)
)

liquido = saldo_bruto - total_deducoes

st.markdown("---")
st.markdown("<div class='section-header'>Resultado</div>", unsafe_allow_html=True)

c1, c2, c3, c4 = st.columns(4)
with c1:
    st.markdown(f"""<div class='result-card'>
        <div class='result-label'>Frete total</div>
        <div class='result-value val-blue'>R$ {frete:,.0f}</div>
    </div>""".replace(",", "X").replace(".", ",").replace("X", "."), unsafe_allow_html=True)

with c2:
    st.markdown(f"""<div class='result-card'>
        <div class='result-label'>Adiantamento</div>
        <div class='result-value val-amber'>R$ {adiantamento:,.0f}</div>
    </div>""".replace(",", "X").replace(".", ",").replace("X", "."), unsafe_allow_html=True)

with c3:
    st.markdown(f"""<div class='result-card'>
        <div class='result-label'>Deduções</div>
        <div class='result-value val-red'>− R$ {total_deducoes:,.0f}</div>
    </div>""".replace(",", "X").replace(".", ",").replace("X", "."), unsafe_allow_html=True)

with c4:
    cor = "val-green" if liquido >= 0 else "val-red"
    st.markdown(f"""<div class='result-card'>
        <div class='result-label'>Líquido motorista</div>
        <div class='result-value {cor}'>R$ {max(0,liquido):,.0f}</div>
    </div>""".replace(",", "X").replace(".", ",").replace("X", "."), unsafe_allow_html=True)

partes = []
if usar_comissao: partes.append(f"comissão R${comissao:,.0f}")
if usar_guarda:   partes.append(f"guarda R${guarda:,.0f}")
if usar_gerente:  partes.append(f"gerente R${gerente:,.0f}")
if usar_pneu:     partes.append(f"pneu/borr. R${pneu:,.0f}")
if usar_outros and outros > 0: partes.append(f"{outros_desc or 'outros'} R${outros:,.0f}")
ded_str = " + ".join(partes) if partes else "sem deduções"

if abastecimento > 0:
    obs_abast = f"  |  abastecimento registrado: R$ {abastecimento:,.0f}"
else:
    obs_abast = ""

formula_html = f"""
<div class='formula-box'>
{motorista or "Motorista"} &nbsp;·&nbsp; {data_viagem.strftime('%d/%m/%Y')} &nbsp;·&nbsp; {rota_selecionada}<br>
R$ {frete:,.0f} × {pct_adi}% = R$ {adiantamento:,.0f} adiantado<br>
saldo bruto R$ {saldo_bruto:,.0f} − [{ded_str}] = <strong style='color:#4ade80'>R$ {max(0,liquido):,.0f} líquido</strong>{obs_abast}
</div>
"""
st.markdown(formula_html.replace(",", "X").replace(".", ",").replace("X", "."), unsafe_allow_html=True)

st.markdown("---")
st.markdown("<div class='section-header'>Exportar</div>", unsafe_allow_html=True)

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Viagem"

    # Estilos
    cor_header   = "1E2433"
    cor_secao    = "252A3D"
    cor_positivo = "166534"
    cor_negativo = "7F1D1D"
    cor_neutro   = "1A1D27"

    fonte_titulo  = Font(name="Arial", bold=True, size=14, color="E8E6E0")
    fonte_secao   = Font(name="Arial", bold=True, size=10, color="9CA3AF")
    fonte_label   = Font(name="Arial", size=10, color="9CA3AF")
    fonte_valor   = Font(name="Arial", bold=True, size=11, color="E8E6E0")
    fonte_verde   = Font(name="Arial", bold=True, size=12, color="4ADE80")
    fonte_vermelha= Font(name="Arial", bold=True, size=11, color="F87171")
    fonte_azul    = Font(name="Arial", bold=True, size=11, color="60A5FA")
    fonte_amber   = Font(name="Arial", bold=True, size=11, color="FBBF24")
    borda_fina    = Border(
        bottom=Side(style="thin", color="2E3347"),
        top=Side(style="thin", color="2E3347"),
    )

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def set_row(row, label, valor, fonte_v=None, bg=None, bold_label=False):
        c_lbl = ws.cell(row=row, column=1, value=label)
        c_lbl.font = Font(name="Arial", bold=bold_label, size=10,
                          color="9CA3AF" if not bold_label else "E8E6E0")
        c_lbl.alignment = Alignment(vertical="center")
        c_val = ws.cell(row=row, column=2, value=valor)
        c_val.font = fonte_v or fonte_valor
        c_val.alignment = Alignment(horizontal="right", vertical="center")
        c_val.number_format = 'R$ #.##0,00'
        if bg:
            c_lbl.fill = fill(bg)
            c_val.fill = fill(bg)
        ws.row_dimensions[row].height = 20
        return c_lbl, c_val

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Título
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "RELATÓRIO DE VIAGEM"
    t.font = fonte_titulo
    t.fill = fill(cor_header)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Cabeçalho info
    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value = f"Gerado em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub.font = Font(name="Arial", size=9, color="4B5563")
    sub.fill = fill(cor_header)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    r = 4
    # Seção: Identificação
    ws.merge_cells(f"A{r}:B{r}")
    s = ws[f"A{r}"]
    s.value = "IDENTIFICAÇÃO"
    s.font = fonte_secao
    s.fill = fill(cor_secao)
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    set_row(r, "Motorista", motorista or "—",
            Font(name="Arial", bold=True, size=11, color="E8E6E0"), cor_neutro)
    r += 1
    c_lbl, c_val = set_row(r, "Data da viagem", None, None, cor_neutro)
    c_val.value = data_viagem
    c_val.number_format = "DD/MM/YYYY"
    c_val.font = fonte_valor
    r += 1
    set_row(r, "Rota", rota_selecionada, fonte_valor, cor_neutro)
    r += 1

    r += 1
    # Seção: Valores
    ws.merge_cells(f"A{r}:B{r}")
    s = ws[f"A{r}"]
    s.value = "VALORES"
    s.font = fonte_secao
    s.fill = fill(cor_secao)
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    set_row(r, "Frete bruto", frete, fonte_azul, cor_neutro)
    r += 1
    set_row(r, f"Adiantamento ({pct_adi}%)", adiantamento, fonte_amber, cor_neutro)
    r += 1
    set_row(r, "Saldo bruto", saldo_bruto, fonte_valor, cor_neutro)
    r += 1
    set_row(r, "Abastecimento", abastecimento, fonte_valor, cor_neutro)
    r += 1

    r += 1
    # Seção: Deduções
    ws.merge_cells(f"A{r}:B{r}")
    s = ws[f"A{r}"]
    s.value = "DEDUÇÕES DO SALDO"
    s.font = fonte_secao
    s.fill = fill(cor_secao)
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    deducoes_itens = [
        ("Comissão",          comissao if usar_comissao else 0),
        ("Guarda",            guarda   if usar_guarda   else 0),
        ("Gerente",           gerente  if usar_gerente  else 0),
        ("Pneu / borracharia",pneu     if usar_pneu     else 0),
        (outros_desc or "Outros", outros if usar_outros else 0),
    ]
    for desc, val in deducoes_itens:
        set_row(r, desc, val, fonte_vermelha if val > 0 else fonte_label, cor_neutro)
        r += 1

    set_row(r, "Total deduções", total_deducoes, fonte_vermelha, cor_neutro, bold_label=True)
    r += 1

    r += 1
    # Resultado final
    ws.merge_cells(f"A{r}:B{r}")
    s = ws[f"A{r}"]
    s.value = "RESULTADO"
    s.font = fonte_secao
    s.fill = fill(cor_secao)
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    cor_res = "166534" if liquido >= 0 else "7F1D1D"
    ws.merge_cells(f"A{r}:B{r}")
    res = ws[f"A{r}"]
    res.value = f"Pagamento líquido: R$ {max(0, liquido):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    res.font = Font(name="Arial", bold=True, size=14, color="4ADE80" if liquido >= 0 else "F87171")
    res.fill = fill(cor_res)
    res.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

nome_xlsx = f"frete_{(motorista or 'motorista').lower().replace(' ','_')}_{data_viagem.strftime('%d%m%Y')}.xlsx"

st.download_button(
    label="⬇ Baixar relatório (.xlsx)",
    data=gerar_xlsx(),
    file_name=nome_xlsx,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)